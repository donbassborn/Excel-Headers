#
#   Boot Loop © 2020
#   Author: x3x3n0m0rph
#

import os
import posixpath
import csv
import openpyxl
import pathlib
import time


class HeaderUnit:
    name = "" 
    id = -1
    pos = -1
    id_parent = -1
    level = -1
    chapter = ""
    year = 2020
    is_actual = ""

    starts = -1
    ends = -1

def deep(sheet, start, end, level, id_parent, start_row, end_row, header_units, id_iterator):
    col = start
    while col <= end:
        unit = HeaderUnit()

        unit.id = id_iterator[0]
        id_iterator[0] += 1
        unit.id_parent = id_parent
        unit.starts = col
        unit.ends = col
        unit.name = str(sheet.cell(start_row+level,col).value).strip()
        unit.level = level
        unit.chapter = sheet.title
        
        
        while(col != end and sheet.cell(start_row+level,col+1).value == None):
            col += 1
            unit.ends = col

        #print("\t"*level, unit.id, "name:", unit.name.replace("\n", " "), "[", unit.starts, "-",  unit.ends, "]", "parent:", unit.id_parent) #debug
        
        if sheet.cell(start_row+level+1,unit.starts).value != None and start_row+level < end_row-1:
            deep(sheet, unit.starts, unit.ends, level+1, unit.id, start_row, end_row, header_units, id_iterator)
            header_units.append(unit)
        else:
            unit.pos = col
            header_units.append(unit)
        col += 1
        



if __name__ == "__main__":
    path = pathlib.Path(__file__).parent.absolute().as_posix()
    csv_dir = "csv"
    csv_delim = ";"
    delim = posixpath.sep
    pattern = "*.xls*"
    file_counter = 0  

    print("Active directory:", path)
    print("Script will process all", pattern, "files")

    files = os.listdir(path)  
    for filename in files:  
        if (filename.endswith(".xls") | filename.endswith(".xlsx")) & (not filename.startswith("~$")):
            header_units = list()
            loadtime = time.perf_counter()
            wb = openpyxl.load_workbook(path + delim + filename)
            loadtime = time.perf_counter() - loadtime
            print("===========================================================================")
            print("Open excel file: " + path + delim + filename)
            print("Loaded in:", round(loadtime, 2), "sec")
            print("Sheets in file: ", len(wb.sheetnames))
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                print("\tActive sheet:", sheet_name)
                found_header = False
                head_ends_row = -1
                for i in range(1, sheet.max_row):
                    #print(sheet.cell(i,1).value)
                    if str(sheet.cell(i,1).value) == "1" or sheet.cell(i,1).value == 1.0:
                        found_header = True
                        head_ends_row = i
                        print("\t\tFound head 'pos' row:", i)
                        break
                right_col = 1
                if found_header:
                    while sheet.cell(head_ends_row, right_col+1).value != None:
                        right_col += 1

                    head_starts_row = -1

                    for k in range(head_ends_row-1, 0, -1):
                        if sheet.cell(k,1).border.top.style == openpyxl.styles.borders.BORDER_MEDIUM:
                            head_starts_row = k
                            print("\t\tHead dimentions: ", head_starts_row, 1, head_ends_row, right_col)
                            break

                    if head_starts_row > -1:
                        id_iterator = [0]
                        deep(sheet, 1, right_col, 0, -1, head_starts_row, head_ends_row, header_units, id_iterator)
                    

                    else:
                        print("\t\tCannot find first line of head!")
                        #TODO exception
                else:
                    print("\t\tCannot find head in sheet, skipped")

            if len(header_units) > 0:
                print("Found", len(header_units), "params, writing in", path + delim + csv_dir + delim + filename + ".csv")
                if not os.path.exists(path + delim + csv_dir):
                    print("\tNot found directory for csv-files, creating...")
                    os.makedirs(path + delim + csv_dir)
                with open(path + delim + csv_dir + delim + filename + ".csv", "w+",newline='') as csv_file:
                    writer = csv.writer(csv_file, delimiter=csv_delim, )
                    writer.writerow([
                        "name",
                        "id",
                        "pos",
                        "id_parent",
                        "level",
                        "chapter",
                        "year",
                        "is_actual"
                    ])
                    for unit in header_units:
                        writer.writerow([
                            unit.name,
                            unit.id,
                            unit.pos if unit.pos != -1 else None,
                            unit.id_parent if unit.id_parent != -1 else None,
                            unit.level,
                            unit.chapter,
                            unit.year,
                            unit.is_actual
                        ])
            else:
                print("Found 0 sheets with heads, csv-file is not created")
            header_units.clear()
            wb.close()
            file_counter += 1
    print("===========================================================================")
    print("Files processed:", file_counter)
        
                    

                            

                    